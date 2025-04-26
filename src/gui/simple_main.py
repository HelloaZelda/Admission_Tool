import os
import csv
import sys
import traceback
import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import xlrd  # 用于读取Excel文件
from openpyxl import Workbook  # 替换xlwt，使用openpyxl
from openpyxl import load_workbook  # 用于读取xlsx格式
from collections import defaultdict

# 设置日志
def setup_logging():
    log_dir = os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'app.log')
    
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

def get_resource_path(relative_path):
    """获取资源文件的绝对路径"""
    try:
        # PyInstaller创建临时文件夹,将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
    except Exception:
        # 如果不是打包的情况,则使用当前文件的目录
        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    
    return os.path.join(base_path, relative_path)

class SimpleMajorAdmissionApp:
    def __init__(self, root):
        try:
            self.root = root
            self.root.title("本科生专业方向录取软件 V1.0")
            self.root.geometry("800x700")  # 增加窗口高度以适应LOGO
            
            # 添加异常处理
            self.root.report_callback_exception = self.handle_exception
            
            # Initialize data
            self.student_data = []
            self.major_quotas = {
                "电子信息工程": tk.IntVar(value=0),
                "通信工程": tk.IntVar(value=0),
                "电磁场与无线技术": tk.IntVar(value=0)
            }
            
            # 志愿映射
            self.preference_mapping = {
                'A': ['电子信息工程', '通信工程', '电磁场与无线技术'],
                'B': ['电子信息工程', '电磁场与无线技术', '通信工程'],
                'C': ['电磁场与无线技术', '电子信息工程', '通信工程'],
                'D': ['电磁场与无线技术', '通信工程', '电子信息工程'],
                'E': ['通信工程', '电子信息工程', '电磁场与无线技术'],
                'F': ['通信工程', '电磁场与无线技术', '电子信息工程']
            }
            
            self.init_ui()
        except Exception as e:
            logging.error(f"初始化失败: {str(e)}")
            logging.error(traceback.format_exc())
            messagebox.showerror("错误", f"程序初始化失败：{str(e)}\n请查看日志文件了解详情。")
            
    def handle_exception(self, exc_type, exc_value, exc_traceback):
        """处理未捕获的异常"""
        error_msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_traceback))
        logging.error(f"未捕获的异常:\n{error_msg}")
        messagebox.showerror("错误", f"发生错误：{str(exc_value)}\n请查看日志文件了解详情。")

    def init_ui(self):
        try:
            # Create main frame
            main_frame = ttk.Frame(self.root, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Add logo
            try:
                # 使用新的资源路径获取函数
                logo_path = get_resource_path(os.path.join('resources', 'logo.png'))
                if not os.path.exists(logo_path):
                    # 如果找不到PNG，尝试加载ICO格式
                    logo_path = get_resource_path(os.path.join('resources', 'logo.ico'))
                
                if os.path.exists(logo_path):
                    logo_img = Image.open(logo_path)
                    # 计算调整后的大小，保持宽高比
                    target_width = 250  # 目标宽度改为250像素
                    width_percent = (target_width / float(logo_img.size[0]))
                    target_height = int((float(logo_img.size[1]) * float(width_percent)))
                    
                    # 调整LOGO大小
                    logo_img = logo_img.resize((target_width, target_height), Image.Resampling.LANCZOS)
                    logo_photo = ImageTk.PhotoImage(logo_img)
                    logo_label = ttk.Label(main_frame, image=logo_photo)
                    logo_label.image = logo_photo  # 保持引用
                    logo_label.pack(pady=5)  # 减小上下边距
                else:
                    logging.warning(f"Logo文件不存在: {logo_path}")
                    # 如果无法加载图片，显示完整的学院名称
                    logo_label = ttk.Label(main_frame, text="电子信息与通信学院", font=("Arial", 16, "bold"))
                    logo_label.pack(pady=5)
            except Exception as e:
                logging.warning(f"加载Logo失败: {str(e)}")
                # 如果无法加载图片，显示完整的学院名称
                logo_label = ttk.Label(main_frame, text="电子信息与通信学院", font=("Arial", 16, "bold"))
                logo_label.pack(pady=5)
            
            # Major quotas input section
            quotas_frame = ttk.LabelFrame(main_frame, text="专业录取名额设置", padding="10")
            quotas_frame.pack(fill=tk.X, pady=10)
            
            quotas_inner_frame = ttk.Frame(quotas_frame)
            quotas_inner_frame.pack(fill=tk.X)
            
            for i, (major, var) in enumerate(self.major_quotas.items()):
                major_frame = ttk.Frame(quotas_inner_frame)
                major_frame.pack(side=tk.LEFT, padx=10, expand=True)
                
                ttk.Label(major_frame, text=major).pack()
                spin_box = ttk.Spinbox(
                    major_frame, 
                    from_=0, 
                    to=1000, 
                    textvariable=var,
                    width=10
                )
                spin_box.pack(pady=5)
            
            # File operations section
            file_operations_frame = ttk.Frame(main_frame)
            file_operations_frame.pack(fill=tk.X, pady=10)
            
            import_btn = ttk.Button(file_operations_frame, text="导入学生志愿", command=self.import_student_data)
            import_btn.pack(side=tk.LEFT, padx=5)
            
            process_btn = ttk.Button(file_operations_frame, text="处理录取", command=self.process_admissions)
            process_btn.pack(side=tk.LEFT, padx=5)
            
            export_btn = ttk.Button(file_operations_frame, text="导出录取结果", command=self.export_results)
            export_btn.pack(side=tk.LEFT, padx=5)
            
            # Results table
            table_frame = ttk.LabelFrame(main_frame, text="录取结果", padding="10")
            table_frame.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Create treeview for results
            self.results_tree = ttk.Treeview(table_frame, columns=("序号", "学号", "姓名", "分数", "志愿选择", "录取专业"), show="headings")
            
            # Define columns
            self.results_tree.heading("序号", text="序号")
            self.results_tree.heading("学号", text="学号")
            self.results_tree.heading("姓名", text="姓名")
            self.results_tree.heading("分数", text="分数")
            self.results_tree.heading("志愿选择", text="志愿选择")
            self.results_tree.heading("录取专业", text="录取专业")
            
            # Set column widths
            self.results_tree.column("序号", width=50)
            self.results_tree.column("学号", width=100)
            self.results_tree.column("姓名", width=100)
            self.results_tree.column("分数", width=80)
            self.results_tree.column("志愿选择", width=80)
            self.results_tree.column("录取专业", width=150)
            
            # Add scrollbar
            scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.results_tree.yview)
            self.results_tree.configure(yscroll=scrollbar.set)
            
            # Pack treeview and scrollbar
            self.results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        except Exception as e:
            logging.error(f"初始化UI失败: {str(e)}")
            logging.error(traceback.format_exc())
            messagebox.showerror("错误", f"初始化UI失败：{str(e)}\n请查看日志文件了解详情。")
    
    def import_student_data(self):
        try:
            file_name = filedialog.askopenfilename(
                title="选择学生志愿文件",
                filetypes=[("Excel Files", "*.xlsx *.xls"), ("CSV Files", "*.csv"), ("All Files", "*.*")]
            )
            
            if file_name:
                self.student_data = []
                
                if file_name.endswith('.csv'):
                    # 使用csv模块读取csv文件
                    with open(file_name, 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            student = {
                                '序号': row['序号'],
                                '学号': row['学号'],
                                '姓名': row['姓名'],
                                '分数': float(row['分数']),
                                '志愿选择': str(row['志愿选择']).upper(),  # 转换为大写
                                '专业': row['专业']
                            }
                            self.student_data.append(student)
                
                elif file_name.endswith('.xlsx'):
                    # 使用openpyxl读取xlsx文件
                    wb = load_workbook(file_name)
                    sheet = wb.active
                    
                    # 读取数据
                    for row in sheet.iter_rows(min_row=2):
                        student = {
                            '序号': row[0].value,
                            '学号': row[1].value,
                            '姓名': row[2].value,
                            '分数': float(row[4].value),
                            '志愿选择': str(row[6].value).upper(),  # 转换为大写
                            '专业': row[7].value
                        }
                        self.student_data.append(student)
                else:
                    # 使用xlrd读取xls文件
                    workbook = xlrd.open_workbook(file_name)
                    sheet = workbook.sheet_by_index(0)
                    
                    # 读取数据
                    for row_idx in range(1, sheet.nrows):
                        student = {
                            '序号': sheet.cell_value(row_idx, 0),
                            '学号': sheet.cell_value(row_idx, 1),
                            '姓名': sheet.cell_value(row_idx, 2),
                            '分数': float(sheet.cell_value(row_idx, 4)),
                            '志愿选择': str(sheet.cell_value(row_idx, 6)).upper(),  # 转换为大写
                            '专业': sheet.cell_value(row_idx, 7)
                        }
                        self.student_data.append(student)
                
                self.update_results_table()
                messagebox.showinfo("成功", f"成功导入 {len(self.student_data)} 条学生数据")
        except Exception as e:
            messagebox.showerror("错误", f"导入文件时发生错误：{str(e)}")
            logging.error(f"导入文件时发生错误: {str(e)}")
            logging.error(traceback.format_exc())
    
    def process_admissions(self):
        if not self.student_data:
            messagebox.showwarning("警告", "请先导入学生数据")
            return
        
        try:
            # 获取当前名额
            quotas = {major: var.get() for major, var in self.major_quotas.items()}
            
            # 检查是否所有专业都设置了名额
            if all(quota == 0 for quota in quotas.values()):
                messagebox.showwarning("警告", "请先设置专业录取名额")
                return
                
            remaining_quotas = quotas.copy()
            
            # 将所有学生按分数排序
            sorted_students = sorted(self.student_data, key=lambda x: float(x['分数']), reverse=True)
            
            # 第一轮：处理所有学生的第一志愿
            for student in sorted_students:
                choice = student['志愿选择']
                if choice not in self.preference_mapping:
                    student['录取专业'] = '无效志愿'
                    continue
                
                # 获取学生的志愿顺序
                preferences = self.preference_mapping[choice]
                first_choice = preferences[0]
                
                # 如果第一志愿还有名额，直接录取
                if remaining_quotas[first_choice] > 0:
                    student['录取专业'] = first_choice
                    remaining_quotas[first_choice] -= 1
            
            # 第二轮：处理未被录取学生的第二志愿
            for student in sorted_students:
                if '录取专业' not in student or student['录取专业'] == '无效志愿':
                    continue
                
                if '录取专业' not in student:
                    choice = student['志愿选择']
                    preferences = self.preference_mapping[choice]
                    second_choice = preferences[1]
                    
                    if remaining_quotas[second_choice] > 0:
                        student['录取专业'] = second_choice
                        remaining_quotas[second_choice] -= 1
            
            # 第三轮：处理未被录取学生的第三志愿
            for student in sorted_students:
                if '录取专业' not in student or student['录取专业'] == '无效志愿':
                    continue
                
                if '录取专业' not in student:
                    choice = student['志愿选择']
                    preferences = self.preference_mapping[choice]
                    third_choice = preferences[2]
                    
                    if remaining_quotas[third_choice] > 0:
                        student['录取专业'] = third_choice
                        remaining_quotas[third_choice] -= 1
            
            # 处理未被录取的学生（调剂）
            for student in sorted_students:
                if '录取专业' not in student or student['录取专业'] == '无效志愿':
                    # 查找还有剩余名额的专业
                    for major, quota in remaining_quotas.items():
                        if quota > 0:
                            student['录取专业'] = f"{major}(调剂)"
                            remaining_quotas[major] -= 1
                            break
                    else:
                        student['录取专业'] = '未分配'
            
            self.update_results_table()
            
            # 统计录取信息
            total_students = len(self.student_data)
            admitted_count = sum(1 for s in self.student_data if '未分配' not in s['录取专业'])
            not_admitted_count = total_students - admitted_count
            
            stats = {
                '电子信息工程': {'total': 0, 'adjust': 0},
                '通信工程': {'total': 0, 'adjust': 0},
                '电磁场与无线技术': {'total': 0, 'adjust': 0},
                '未分配': 0
            }
            
            for student in self.student_data:
                major = student['录取专业']
                if '未分配' in major:
                    stats['未分配'] += 1
                else:
                    base_major = major.replace('(调剂)', '')
                    if base_major in stats:
                        stats[base_major]['total'] += 1
                        if '调剂' in major:
                            stats[base_major]['adjust'] += 1
            
            # 生成详细的统计信息
            result_msg = "录取完成！\n\n"
            result_msg += f"总人数：{total_students}人\n"
            result_msg += f"已录取：{admitted_count}人\n"
            result_msg += f"未录取：{not_admitted_count}人\n\n"
            result_msg += "各专业录取情况：\n"
            
            for major, data in stats.items():
                if major != '未分配':
                    total = data['total']
                    adjust = data['adjust']
                    normal = total - adjust
                    result_msg += f"\n{major}：\n"
                    result_msg += f"  - 总计：{total}人\n"
                    result_msg += f"  - 正常录取：{normal}人\n"
                    result_msg += f"  - 调剂录取：{adjust}人\n"
                    result_msg += f"  - 剩余名额：{remaining_quotas[major]}人\n"
            
            result_msg += f"\n未分配人数：{stats['未分配']}人"
            
            messagebox.showinfo("录取完成", result_msg)
            
        except Exception as e:
            messagebox.showerror("错误", f"处理录取时发生错误：{str(e)}")
            import traceback
            print(traceback.format_exc())
    
    def export_results(self):
        if not self.student_data:
            messagebox.showwarning("警告", "请先导入学生数据")
            return
            
        try:
            file_name = filedialog.asksaveasfilename(
                title="保存录取结果",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
            )
            
            if file_name:
                # 创建新的工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = '录取结果'
                
                # 写入表头
                headers = ['序号', '学号', '姓名', '分数', '志愿选择', '录取专业']
                ws.append(headers)
                
                # 写入数据
                for student in self.student_data:
                    ws.append([
                        student['序号'],
                        student['学号'],
                        student['姓名'],
                        student['分数'],
                        student['志愿选择'],
                        student.get('录取专业', '')
                    ])
                
                # 调整列宽
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # 保存文件
                wb.save(file_name)
                
                messagebox.showinfo("成功", "录取结果已成功导出")
                
                # 询问是否打开文件
                if messagebox.askyesno("确认", "是否立即打开导出的文件？"):
                    os.startfile(file_name)
                
        except Exception as e:
            messagebox.showerror("错误", f"导出文件时发生错误：{str(e)}")
            logging.error(f"导出文件时发生错误: {str(e)}")
            logging.error(traceback.format_exc())
    
    def update_results_table(self):
        # 清除现有项目
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        # 添加数据到树形视图
        for student in self.student_data:
            self.results_tree.insert(
                "", 
                tk.END, 
                values=(
                    student['序号'],
                    student['学号'],
                    student['姓名'],
                    student['分数'],
                    student['志愿选择'],
                    student.get('录取专业', '')
                )
            )

def main():
    try:
        # 设置日志
        setup_logging()
        logging.info("程序启动")
        
        # 创建主窗口
        root = tk.Tk()
        
        # 设置窗口图标
        try:
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            icon_path = os.path.join(base_path, "..", "..", "resources", "logo.ico")
            if os.path.exists(icon_path):
                root.iconbitmap(icon_path)
        except Exception as e:
            logging.warning(f"设置窗口图标失败: {str(e)}")
        
        app = SimpleMajorAdmissionApp(root)
        root.mainloop()
        
    except Exception as e:
        logging.error(f"程序运行失败: {str(e)}")
        logging.error(traceback.format_exc())
        messagebox.showerror("错误", f"程序运行失败：{str(e)}\n请查看日志文件了解详情。")
        sys.exit(1)

if __name__ == "__main__":
    main() 